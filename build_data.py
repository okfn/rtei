#!/usr/bin/env python
'''
Data script for the RTEI website.

Check ./build_data.py -h for details
'''
import re
import os
import simplejson as json
import csv
import argparse
import random
import string
from collections import OrderedDict
from decimal import Decimal

from openpyxl import load_workbook

# Change as appropiate
INPUT_FILE = 'rtei/static/data/rtei_data_2017.xlsx'
OUTPUT_DIR = 'rtei/static/data/2017'

COUNTRIES_FILE = 'data/countries.json'

CORE_SHEET = 'All Questionnaires'
MAIN_SCORES_SHEET = 'Country Comparisons'
THEMES_SHEET = 'Cross-cutting Themes'
THEMES_MAPPINGS_SHEET = 'Transversal Themes Mappings'

INSUFFICIENT_DATA = 'Insufficient data'


# Don't include this in the parsed data
EXCLUDE_THEMES = ['10', '10A', '10B', '9B']


MODIFIERS = {
    'gp': 'Gender Parity',
    'ad': 'Advantaged Group',
    'resp': 'Residential Parity',
    'disp': 'Disability Parity',
    # These are underscores in the spreadsheet
    'inc-hmp': 'High to Medium Quartile Income Ratio',
    'inc-mlp': 'Medium to Low Quartile Income Ratio',
    'Overage': 'Overage Learners',
    'Out': 'Out-of-school Rate',

}

INDICATOR_TYPES = {
    '3.3.1': {
      'a': 'Primary Schools',
      'b': 'Secondary Schools',
      'c': 'TVET',
      'd': 'Tertiary',
    },
    '3.3.2': {
      'a': 'Primary Schools',
      'b': 'Secondary Schools',
      'c': 'TVET',
      'd': 'Tertiary',
    },
    '3.3.3': {
      'a': 'Primary Schools',
      'b': 'Secondary Schools',
      'c': 'TVET',
      'd': 'Tertiary',
    },
    '4.3.3': {
      'a': 'Overall Primary Schools',
      'b': 'Reading Primary Schools',
      'c': 'Math Primary Schools',
      'd': 'Overall Secondary Schools',
      'e': 'Reading Secondary Schools',
      'f': 'Math Secondary Schools',
    },
    '4.3.4': {
        'a': 'Youth',
        'b': 'Adult',
    }
}

MAIN_INDICATORS = [
    {
        'code': '1',
        'title': 'Governance',
        'level': 1
    },
    {
        'code': '2',
        'title': 'Availability',
        'level': 1
    },
    {
        'code': '3',
        'title': 'Accessibility',
        'level': 1
    },
    {
        'code': '4',
        'title': 'Acceptability',
        'level': 1
    },
    {
        'code': '5',
        'title': 'Adaptability',
        'level': 1
    },
]


EXTRA_INDICATORS = [
    {
        'code': 'S',
        'title': 'Structure',
        'level': 1,
        'show_in_menu': False,
    },
    {
        'code': 'P',
        'title': 'Process',
        'level': 1,
        'show_in_menu': False,
    },
    {
        'code': 'O',
        'title': 'Outcome',
        'level': 1,
        'show_in_menu': False,
    },
]

# Excel handler
wb = None

# Countries dictionary
countries = None

'''
This will match any codes at the beginning of the string, starting with
a number and ending with a space, eg:
1
1.1
1.1.1aa
1.1.1a_emp_dis
'''
valid_code = re.compile(r'^(C )?(\d)(.*?)? ')


def get_country_code(country_name, code_type='iso2'):
    '''
    Given a country name, return its ISO code

    By default to the 2 digit code is returned (eg 'AF' for Afghanistan),
    passing code_type='iso3' will return the 3 digit code (eg'AFG').
    '''
    if not country_name:
        return None
    for code, country in countries.iteritems():
        if (country_name.lower() == country['name'].lower() or
                country_name.lower() in [x.lower() for x in country.get('other_names', [])]):
            return country[code_type]

    return None


def get_country_name(country_code):
    '''
    Given a country code, return the corresponding country name.
    '''
    code_type = 'iso2'
    if len(country_code) == 3:
        code_type = 'iso3'

    return next(
        (c['name'] for code, c in countries.iteritems()
         if c[code_type] == country_code.upper()),
        None)


def get_numeric_cell_value(cell):
    '''
    Return a numeric value rounded to 2 places if it is a float, or the value
    otherwise
    '''
    if isinstance(cell.value, float):
        return Decimal(cell.value)
    else:
        return cell.value


def get_level_for_indicator(code):

    if code.count('.') == 0:
        # 1
        level = 1
    elif code.count('.') == 1:
        # 1.2
        level = 2
    elif code.count('.') == 2:
        # 1.3.4
        level = 3

    elif code.count('.') == 3:
            # 1.5.6a or 1.5.6a_dis
        level = 4
    else:
        raise ValueError('Wrong indicator code: {0}'.format(code))

    return level


def get_level_for_theme(code):
    if not code[-1].isalpha():
        # 1
        level = 1
    else:
        # 1A or 1A.B
        level = 2
    return level


def parse_cell(value, theme=False):
    '''
    Parses an indicator cell to return a tuple with its code, title and level

    Indicator cells must have one of the following formats:

    * {code}: {title}
    * {code}

    Codes must start with a number but they can be pretty much anything else
    after it.
    If there is no title available, for a limited set of conditions it will be
    generated (codes ending in `_year` and those ending in one of the
    MODIFIERS)

    The level is calculated from the code, checking for the number of `.`
    characters and extra modifiers.

    Returns a tuple (code, title, level), with any of the values set to None if
    it could not be computed (which should not happen).
    '''

    code = None
    title = None
    level = None

    match = valid_code.match(value)

    if match:
        groups = match.groups()
        code = ''.join(g for g in groups if g and g != 'C ')
        code = code.rstrip(':').rstrip('.')
        title = value \
            .replace(code + ':', '') \
            .replace('C ', '') \
            .replace('\u00a0', '') \
            .replace(code, '') \
            .strip('.') \
            .strip()
        level = (get_level_for_theme(code) if theme
                 else get_level_for_indicator(code))
    elif '_year' in value.split()[0]:
        code = value.split()[0]
        title = 'Year'
        level = 4
    elif any('_' + modifier.replace('-', '_') in value.split()[0]
             for modifier in MODIFIERS.keys()):
        # This is a derived indicator, we compute the title automatically
        # in the form:
        # Indicator type: Modifer - Modifier
        # from a code like 4.3.3a_gp_disp
        parts = value.replace('inc_', 'inc-').split('_')
        title_parts = []
        for part in parts[1:]:
            title_parts.append(MODIFIERS[part])
        code = value[:5]
        indicator_type = INDICATOR_TYPES.get(code, {}).get(parts[0][-1], '')
        if indicator_type:
            title = '{0}: {1}'.format(indicator_type, ' - '.join(title_parts))
        else:
            print 'Warning, could not get indicator type for {0}'.format(value)
            title = ' - '.join(title_parts)

        code = value.split()[0]
        level = 4

    # Replace fake apostrophes
    if title:
        title = title.replace(u'\u2019', '\'')

    return code, title, level


def get_themes(include_rows=False, include_indicators=False):
    '''
    Returns a list of dicts describing the transversal themes. Each of the
    dicts contains the following keys:

        * `code`: The unique identifier for this indicator, examples include
            1, 1.A, 1.A.B
        * `title`: The string that describes the indicator.
        * `level`: The hierarchy of the indicator (level 1 or 2)
        * `row` (optional, only if `include_rows` is True): The row on
            the spreadsheet that holds the values for the indicator.

    '''

    out = []
    sheet = THEMES_SHEET
    ws = wb[sheet]

    codes_done = []
    theme = None
    for i in xrange(2, len(ws.rows) + 1):
        cell = ws.cell(row=i, column=1)
        if not cell.value:
            continue

        code, title, level = parse_cell(cell.value, theme=True)
        if not code:
            # Assume it's a computed indicator
            level = 4
            title = cell.value
        elif len(code.split('.')) == 3:
            level = 4

        if code in EXCLUDE_THEMES:
            continue

        if level in (1, 2) and code in codes_done:
            if include_rows:
                previous_theme = [t for t in out if t['code'] == code][0]
                # This is the values row
                previous_theme['row'] = i
            continue

        codes_done.append(code)

        if level in (1, 2):
            theme = {
                'code': code,
                'title': title,
                'level': level
            }

            out.append(theme)

        if include_indicators and level in (3, 4):
            if 'indicators' not in theme:
                theme['indicators'] = []

            theme['indicators'].append({
                'code': code,
                'title': title
            })

    return out


def get_indicators(include_columns=False):
    '''
    Returns a list of dicts describing the indicators. Each of the dicts
    contains the following keys:

        * `code`: The unique identifier for this indicator, examples include
            1, 2.3, 3.3.4, 1.3a, 2.2.1_year, 3.2.1a_resp_ad.
        * `title`: The string that describes the indicator.
        * `level`: The hierarchy of the indicator. Without taking the derived
            indicators into account this is straight forward: 1 -> 1, 1.1 -> 2,
            1.1.1 -> 3, 1.1.1a -> 4. Derived indicators can have the following
            levels: 1.2a -> 2 and 1.1.1a_resp_ad -> 4.
        * `column` (optional, only if `include_columns` is True): The column on
            the spreadsheet that holds the values for the indicator.

    '''

    out = []

    ws = wb[CORE_SHEET]

    # First four rows
    codes_done = []
    for i in (0, 1, 2, 3):
        for cell in ws.rows[i]:
            indicator = None
            if cell.value:
                code, title, level = parse_cell(cell.value)

                if not code:
                    continue
                if code in codes_done:
                    if i == 3:
                        # Use the previous one
                        indicator = [o for o in out if o['code'] == code][0]
                    else:
                        continue
                elif '_year' in code and i == 3:
                    indicator = [o for o in out
                                 if o['code'] == code.replace('_year', '')][0]
                else:
                    codes_done.append(code)

                if not indicator:
                    indicator = {
                        'code': code,
                        'title': title,
                        'level': level
                    }
                    out.append(indicator)
                if i == 3 and include_columns:
                    if title == 'Response':
                        indicator['column_response'] = cell.column
                    elif '_year' in code:
                        indicator['column_year'] = cell.column
                    else:
                        indicator['column'] = cell.column
            else:
                # Other indicators
                # print cell.value
                pass
    return out


def get_all_indicators(include_columns=False):
    '''
    Returns a list of all indicators. See get_indicators for details.
    '''

    indicators = get_indicators(include_columns=include_columns)

    # Add extra indicators not defined in the core sheet
    indicators = indicators + EXTRA_INDICATORS

    return indicators


def get_main_scores():
    '''
    Return the values for the level 1 indicators (ie 1, 2, 3, 4 and 5),
    SPO and global index for a particular country.

    These are read directly from the MAIN_SCORES_SHEET

    '''

    ws = wb[MAIN_SCORES_SHEET]

    row_index = None

    value_columns = [chr(i) for i in range(ord('B'), ord('J') + 1)]

    indicators = MAIN_INDICATORS + EXTRA_INDICATORS

    out = {}

    # Get the country row
    for cell in ws.columns[0]:

        name = cell.value

        if not name:
            if cell.row > 3:
                break
            else:
                continue

        if name.lower() == 'country':
            continue

        country_code = get_country_code(name)
        if not country_code:
            print('Could not find a data row for country {}'.format(country_code))
            continue

        row_index = cell.row

        out[country_code] = {}

        # Read values for that row (columns B to J)
        for column in value_columns:
            indicator_cell = '{}{}'.format(column, 2)
            indicator_title = ws[indicator_cell].value

            # Get indicator code
            if indicator_title.lower() == 'index score':
                indicator_code = 'index'
            else:
                for indicator in indicators:
                    if indicator_title.lower() == indicator['title'].lower():
                        indicator_code = indicator['code']
                        break
            if not indicator_code:
                print('Could not find code for indicator {}'.format(indicator_title))
                continue

            value_cell = '{}{}'.format(column, row_index)
            value = get_numeric_cell_value(ws[value_cell])

            out[country_code][indicator_code] = value * 100 if value <= 1 else value

    return out

def add_main_scores(indicators):
    '''
    Add the values for the level 1 indicators (ie 1, 2, 3, 4 and 5) to the
    passed country_indicators dict.

    '''
    main_scores = get_main_scores()
    for k in indicators:
        if k not in main_scores:
            continue
        indicators[k].update(main_scores[k])


def indicators_per_country(max_level=4, derived=True, random_values=False,
                           responses=True):
    '''
    Returns a dict containing the values for all indicators for all countries,
    with the following structure:

        {
            'CL': {
                    '1': 64.76,
                    '1.1': 100.0,
                    '1.1.1a': 'Yes',
                    '1.1.1b': 'Yes',
                    '1.1.1c': 'Yes',
                    # ...
                    },
            'TZ': {
                    '1': 74.736,
                    '1.1': 100.0,
                    '1.1.1a': 'Yes',
                    '1.1.1b': 'No',
                    '1.1.1c': 'Yes',
                    # ...
                    },

            # ...
        }

    If responses is False, the actual indicator value is returned:
        {
            'CL': {
                    '1': 64.76,
                    '1.1': 100.0,
                    '1.1.1a': 1,
                    '1.1.1b': 1,
                    '1.1.1c': 1,
                    # ...
                    },
            'TZ': {
                    '1': 74.736,
                    '1.1': 100.0,
                    '1.1.1a': 1,
                    '1.1.1b': 0,
                    '1.1.1c': 1,
                    # ...
                    },

            # ...
        }

    You can restrict the indicators returned with the `max_level` parameter
    and whether to return `derived` (eg 1.2a or 3.2.1_ag) indicators or not.

    If `random_values` is True, countries not present in the spreadsheet are
    returned with random values (only for levels 1 and 2).

    Values with `999` are replaced with `No data`.

    '''

    out = OrderedDict()

    indicators = get_all_indicators(include_columns=True)

    ws_core = wb[CORE_SHEET]
    country_codes = []

    for i in xrange(5, len(ws_core.rows) + 1):
        country_name = ws_core['A' + str(i)].value
        if not country_name:
            continue
        country_code = get_country_code(country_name)
        if not country_code:
            print 'Warning: Could not get country code for {0}'.format(
                country_name)
        country_codes.append(country_code)
        out[country_code] = OrderedDict()

    for i, country_code in enumerate(country_codes):
        for indicator in indicators:
            if indicator['level'] <= max_level and indicator.get('column'):

                if not derived and indicator['code'][-1].isalpha():
                    # Avoid derived indicators (eg 1.2a or 3.2.1_ag)
                    continue
                if responses and indicator.get('column_response'):
                    cell = indicator['column_response'] + str(i + 5)
                else:
                    cell = indicator['column'] + str(i + 5)

                value = get_numeric_cell_value(ws_core[cell])

                if value == 999:
                    value = 'No data'

                if (responses and indicator.get('column_year')
                        and value not in ('No data', INSUFFICIENT_DATA)):
                    cell_year = indicator['column_year'] + str(i + 5)
                    value = '{0} ({1})'.format(
                        round(value, 2), ws_core[cell_year].value)

                out[country_code][indicator['code']] = value

                # Custom stuff :|

                # Show all level 2, non derived scores (eg 1.2, 3.4, 5.1)
                # as percentages, rounded to 2 places

                if (indicator['code'].count('.') == 1 and
                        not indicator['code'][-1].isalpha() and
                        value is not None and value not in ('No data', INSUFFICIENT_DATA)):
                    out[country_code][indicator['code']] = Decimal(
                        value * 100 if value <= 1 else value)
    add_main_scores(out)

    if random_values:
        for code, country in countries.iteritems():
            if country['iso2'] not in country_codes and country['iso2']:
                out[country['iso2']] = OrderedDict()
                for indicator in indicators:
                    if (indicator['level'] <= 2 and
                            not indicator['code'][-1].isalpha()):
                        out[country['iso2']][indicator['code']] = \
                            random.randint(0, 100)
    return out


def themes_per_country(prefix=None, random_values=False):
    '''
    Returns a dict containing the values for all themes for all countries,
    with the following structure:

        {
            'CL': {
                    '1A.A': 68.4,
                    '2A': 100.0,
                    '3A': 83.07,
                    # ...
                    },
            'TZ': {
                    '1A.A': 83.25,
                    '2A': 100.0,
                    '3A': 63.18,
                    # ...
                    },

            # ...
        }

    If `prefix` is provided, it will be added at the beginning of the code
    (eg to mix themes and indicators).

    If `random_values` is True, countries not present in the spreadsheet are
    returned with random values.
    '''

    out = OrderedDict()

    themes = get_themes(include_rows=True)
    ws = wb[THEMES_SHEET]
    available_countries = []
    for letter in string.ascii_uppercase[1:]:
        country_name = ws[letter + '1'].value
        if not country_name:
            break
        country_code = get_country_code(country_name)
        if not country_code:
            print 'Warning: Could not get country code for {0}'.format(
                country_name)
        available_countries.append((country_code, letter))
        out[country_code] = OrderedDict()

    for i, available_country in enumerate(available_countries):
        country_code, letter = available_country
        for theme in themes:
            if theme['level'] in (1, 2) and theme.get('row'):
                cell = letter + str(theme['row'])

                value = get_numeric_cell_value(ws[cell])

                if isinstance(value, basestring):
                    value = None

                key = str(prefix) + theme['code'] if prefix else theme['code']

#                if isinstance(value, (long, float, Decimal)):
#                    value = round(value, 2)

                # Output as percentages
                if value:
                    value = value * 100

                out[country_code][key] = value

    if random_values:
        country_codes = [c[0] for c in available_countries]
        for code, country in countries.iteritems():
            if country['iso2'] not in country_codes and country['iso2']:
                out[country['iso2']] = OrderedDict()
                for theme in themes:
                    if (theme['level'] == 2):
                        key = (str(prefix) + theme['code'] if prefix
                               else theme['code'])
                        out[country['iso2']][key] = random.randint(
                            0, 100)

    return out


def indicators_as_csv(output_dir=OUTPUT_DIR):
    '''
    Write a CSV file with all indicators, in the form:

        code,title,core,level
        1,Governance,True,1
        1.1,International Framework,True,2
        1.1.1,Is the State party to the United Nations treaties?,True,3

        ...
    '''
    indicators = get_all_indicators()
    output_file = os.path.join(output_dir, 'indicators.csv')

    with open(output_file, 'w') as f:
        w = csv.DictWriter(f, ['code', 'title', 'core', 'level'])
        w.writeheader()
        w.writerows(indicators)


def get_nested_indicators(parent_indicator, all_indicators):
    if parent_indicator['level'] >= 4:
        return
    out = []
    for indicator in all_indicators:
        if (indicator['code'].startswith(parent_indicator['code']) and
                indicator['level'] == parent_indicator['level'] + 1):
            if indicator['level'] < 4:
                nested_indicators = get_nested_indicators(indicator,
                                                          all_indicators)
                if nested_indicators:
                    indicator['children'] = nested_indicators
            out.append(indicator)

    return out


def get_nested_themes(parent_theme, all_themes):
    out = []
    for theme in all_themes:
        numeric_part = ''.join([x for x in theme['code'] if x.isdigit()])
        if (numeric_part == parent_theme['code']
                and theme['level'] == parent_theme['level'] + 1):
            if theme['level'] < 4:
                nested_themes = get_nested_themes(theme,
                                                  all_themes)
                if nested_themes:
                    theme['children'] = nested_themes
            out.append(theme)

    return out


def indicators_as_json(output_dir=OUTPUT_DIR):
    '''
    Write a JSON file with all indicators nested, in the form:

        [
            {
                "code": "1",
                "core": true,
                "level": 1,
                "title": "Governance"
                "children": [
                    {
                        "code": "1.1",
                        "core": true,
                        "level": 2,
                        "title": "International Framework"
                        "children": [
                            {
                                "code": "1.1.1",
                                "core": true,
                                "level": 3,
                                "title": "Is the State party to the [...]"
                            }
                        ]
                    }
                ]
            },

            ...
        ]
    '''

    out = []
    indicators = get_all_indicators()
    for indicator in [i for i in indicators if i['level'] == 1]:
        indicator['children'] = get_nested_indicators(indicator, indicators)
        out.append(indicator)

    output_file = os.path.join(output_dir, 'indicators.json')

    with open(output_file, 'w') as f:
        f.write(json.dumps(out))


def themes_as_json(output_dir=OUTPUT_DIR):
    '''
    Write a JSON file with all themes nested, in the form:

        [
            {
                "code": "1",
                "level": 1,
                "title": "Children with Disabilities"
                "children": [
                    {
                        "code": "1.A.A",
                        "level": 2,
                        "title": "Structure and Support"
                        "indicators": [
                            {
                                "code": "1.4.4j",
                                "title": "Is data disaggregated by [...]"
                            },
                            {
                                "code": "3.2.1j",
                                "title": "Do national laws forbid [...]"
                            },
                            ...
                        ]
                    }
                ]
            },

            ...
        ]
    '''

    out = []
    themes = get_themes(include_indicators=True)

    for theme in [i for i in themes if i['level'] == 1]:
        theme['children'] = get_nested_themes(theme, themes)
        out.append(theme)

    output_file = os.path.join(output_dir, 'themes.json')

    with open(output_file, 'w') as f:
        f.write(json.dumps(out))


def indicators_per_country_as_json(one_file=True, output_dir=OUTPUT_DIR):

    out = indicators_per_country()
    if one_file:
        output_file = os.path.join(output_dir, 'indicators_per_country.json')

        with open(output_file, 'w') as f:
            f.write(json.dumps(out))
    else:
        for country_code in out.keys():
            output_file = os.path.join(output_dir,
                                       '{0}.json'.format(country_code))

            with open(output_file, 'w') as f:
                f.write(json.dumps(out[country_code]))


def scores_per_country_as_json(output_dir=OUTPUT_DIR, random_values=False):

    out = indicators_per_country(max_level=2, derived=False,
                                 random_values=random_values)
    add_main_scores(out)

    file_name = ('scores_per_country.json' if not random_values
                 else 'scores_per_country_random.json')
    output_file = os.path.join(output_dir, file_name)

    with open(output_file, 'w') as f:
        f.write(json.dumps(out))


def scores_per_country_as_csv(output_dir=OUTPUT_DIR, random_values=False):
    '''
    Write a CSV file with the level 1 and 2 scores for each country, as well as
    the transversal themes, in the form:

        iso2,index,1,1.1,1.2,..,t1.A.A
        CL,68.78,100,56.334,89.322,...,68.4

        ...

    Note that transversal theme codes are prefixed with `t` to avoid conflicts
    '''
    out = indicators_per_country(max_level=2, derived=False,
                                 random_values=random_values)
    themes = themes_per_country(prefix='t', random_values=random_values)

    out_list = []
    for country in out.keys():
        if country in themes:
            out[country].update(themes[country])
        # TODO
        #add_main_scores(country, out[country])
        out[country]['iso2'] = country

        out_list.append(out[country])

    file_name = ('scores_per_country.csv' if not random_values
                 else 'scores_per_country_random.csv')
    output_file = os.path.join(output_dir, file_name)

    with open(output_file, 'w') as f:
        w = csv.DictWriter(f, out_list[0].keys())
        w.writeheader()
        w.writerows(out_list)


def c3_ready_json(output_dir=OUTPUT_DIR, random_values=False):
    '''
    Writes a C3 optimized JSON file to display the country bar charts.

    Contains a list of objects, one for each country. First and second level
    indicators are normalized to fit the total percentage of each category.
    Transversal themes are included.

    [
      {
        "1": 12.85,
        "1.1": 20.0,
        "1.2": 10.0,
        "1.3": 0.0,
        "1.4": 16.88,
        "1.5": 17.39,
        "2": 15.13,
        "2.1": 8.32,
        "2.2": 24.0,
        "2.3": 18.3,
        "2.4": 25.0,
        "index": 72.90,
        "name": "Chile",
        "t1A.A": 68.4,
        "t2A": 100,
        "t3A": 83.07,
        "t3B": 72.03,
        "t4A": 55.53,
        "t5A": 76.14,
        "t6A": 71.1,
        "t7A": 75,
        "t7B": 75.59,
        "t8A": 66.66,
        "t9A": 0
      },
      ...
    ]
    '''

    indicators = indicators_per_country(max_level=2, derived=False,
                                        random_values=random_values)

    themes = themes_per_country(prefix='t', random_values=random_values)

    out = []
    for country_code, values in indicators.iteritems():
        item = OrderedDict()

        # Normalize scores
        scores = {str(n): [] for n in xrange(1, 6)}
        scores['main'] = []
        for code, value in values.iteritems():
            if code == 'index':
                continue
            if '.' not in code:
                scores['main'].append(value)
            elif value is not INSUFFICIENT_DATA:
                scores[code[:1]].append(value)
        for code, value in values.iteritems():
            if code == 'index':
                continue
            if code in ('S', 'P', 'O'):
                item[code] = value
            elif '.' not in code:
                if value == INSUFFICIENT_DATA:
                    item[code] = value
                else:
                    item[code] = value / Decimal(len(scores['main']))
            else:
                if value == INSUFFICIENT_DATA:
                    item[code] = value
                else:
                    item[code] = value / Decimal(len(scores[code[:1]]))

        # Add Transversal themes
        item.update(themes[country_code])

        # Add main index
        item['index'] = values['index']

        # Add country name
        item['name'] = get_country_name(country_code)

        out.append(item)

    file_name = ('c3_scores_per_country.json' if not random_values
                 else 'c3_scores_per_country_random.json')
    output_file = os.path.join(output_dir, file_name)

    with open(output_file, 'w') as f:
        f.write(json.dumps(out))


def translation_strings():
    '''
    Outputs all translatable strings in a python module wrapped in the gettext
    function (_), so they can be compiled to the i18n *.po files, eg:

        from django.utils.translation import ugettext as _

        _('Governance')
        _('Accessibility')
        _('Yes')
        _('Missing')
        ...

    The output file is always `rtei/translation_strings.py`
    '''
    indicators = get_indicators()
    themes = get_themes()
    responses = indicators_per_country()

    output_file = os.path.join(os.path.dirname(__file__), 'rtei',
                               'translation_strings.py')

    out = []
    out.append('# This is an auto-generated file via '
               '`build_data.py translation-strings`. ')
    out.append('# DO NOT edit this file directly')
    out.append('from django.utils.translation import ugettext as _')
    out.append('\n\n')

    def translatable_string(string):
        string = string.replace('\'', '\\\'')
        string = string.encode('utf8')
        return'_(\'%s\')' % string

    out.append('# Indicator titles')
    for indicator in indicators:
        out.append(translatable_string(indicator['title']))

    out.append('# Theme titles and custom mapping titles')
    for theme in themes:
        out.append(translatable_string(theme['title']))
        for indicator in theme.get('indicators', []):
            if translatable_string(indicator['title']) not in out:
                out.append(translatable_string(indicator['title']))

    out.append('# Responses')
    responses_done = []
    year = re.compile(r'.*\(20.*\)')
    for country, response_values in responses.iteritems():
        for key, value in response_values.iteritems():
            if (value not in responses_done and
                    isinstance(value, basestring) and
                    not year.match(value)):
                out.append(translatable_string(value))
                responses_done.append(value)

    with open(output_file, 'w') as f:
        f.write('\n'.join(out))


def countries_with_data(output_dir=OUTPUT_DIR):

    data_available = indicators_per_country(max_level=2, derived=False)

    out = {}
    for country_code in data_available.keys():
        out[country_code] = get_country_name(country_code)

    output_file = os.path.join(output_dir, 'countries_with_data.json')

    with open(output_file, 'w') as f:
        f.write(json.dumps(out))


if __name__ == '__main__':

    description = '''
Builds the data files needed for powering the visualizations for the Right to
Education Index website.

The source of data is an Excel file (xlsx). By default this is assumed to be
at `data/rtei_data_2015.xlsx` but you can provide a different location with the
`-i` flag.

The outputs are JSON files, by default created on the `rtei/static/data`
folder. You can change the destination folder with the `-o` flag.

The available outputs are:

    * `indicators-json`
    * `indicators-csv `
    * `themes-json`
    * `scores-per-country-json`
    * `scores-per-country-csv`
    * `indicators-per-country`
    * `c3-ready-json`
    * `translation-strings`
    * `all`
    '''

    parser = argparse.ArgumentParser(
        description=description,
        formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('type',
                        help='Output type')
    parser.add_argument('-i', '--input',
                        help='Input Excel (xlsx) file')
    parser.add_argument('-o', '--output',
                        help='Output directory for the files')
    parser.add_argument('-r', '--random',
                        action='store_true',
                        help='Fill values for missing countries with '
                             'random values')

    args = parser.parse_args()

    input_file = args.input or INPUT_FILE
    output_dir = args.output or OUTPUT_DIR

    wb = load_workbook(input_file, data_only=True)

    with open(COUNTRIES_FILE, 'r') as f:
        countries = json.load(f)
    if args.type == 'all':
        indicators_as_json(output_dir)
        themes_as_json(output_dir)
        indicators_per_country_as_json(one_file=False, output_dir=output_dir)
        scores_per_country_as_json(output_dir, random_values=args.random)
        c3_ready_json(output_dir=output_dir)
        countries_with_data(output_dir=output_dir)
    elif args.type == 'indicators-json':
        indicators_as_json(output_dir)
    elif args.type == 'themes-json':
        themes_as_json(output_dir)
    elif args.type == 'indicators-csv':
        indicators_as_csv(output_dir)
    elif args.type == 'scores-per-country-json':
        scores_per_country_as_json(output_dir, random_values=args.random)
    elif args.type == 'scores-per-country-csv':
        scores_per_country_as_csv(output_dir, random_values=args.random)
    elif args.type == 'indicators-per-country':
        indicators_per_country_as_json(one_file=False, output_dir=output_dir)
    elif args.type == 'c3-ready-json':
        c3_ready_json(output_dir=output_dir, random_values=args.random)
    elif args.type == 'countries-with-data':
        countries_with_data(output_dir=output_dir)
    elif args.type == 'translation-strings':
        translation_strings()
    else:
        print 'Unknown output type'
